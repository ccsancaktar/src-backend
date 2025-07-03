from fastapi import FastAPI, Depends, HTTPException, status, Query, Body
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel
from datetime import datetime, timedelta
from typing import Optional, List, Union, Any
import jwt
from passlib.context import CryptContext
import sqlite3
from dotenv import load_dotenv
import os
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from fastapi.requests import Request
import bcrypt
from fastapi import UploadFile, File, Form
import shutil
from fastapi.staticfiles import StaticFiles
from fastapi import APIRouter
import openpyxl
from tempfile import NamedTemporaryFile

# Ortam değişkenlerini yükle
load_dotenv()

# FastAPI uygulaması oluştur
app = FastAPI()

# uploads klasörünü statik olarak sun
uploads_dir = os.path.join(os.getcwd(), "uploads")
os.makedirs(uploads_dir, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=uploads_dir), name="uploads")

# CORS ayarları
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Production'da spesifik domainler ekleyin
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# JWT Config
SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key-here")
ALGORITHM = os.getenv("ALGORITHM", "HS256")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", 30))

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# SQLite database connection
def get_db_connection():
    try:
        conn = sqlite3.connect('bandirma.db')
        conn.row_factory = sqlite3.Row  # Sözlük benzeri erişim için
        return conn
    except sqlite3.Error as e:
        print(f"SQLite bağlantı hatası: {e}")
        return None

# Modeller
class Token(BaseModel):
    access_token: str
    token_type: str

class User(BaseModel):
    id: int
    username: str
    role: str

class PasswordChangeRequest(BaseModel):
    old_password: str
    new_password: str

class UserCreateRequest(BaseModel):
    username: str
    password: str
    role: str

class KursiyerEkleRequest(BaseModel):
    kurs_durumu: str = ""
    aday_ismi: str = ""
    tel_no: str = ""
    tel_yakin: str = ""
    tc_kimlik: str = ""
    ogrenim_belgesi: Union[int, str] = 0
    adres_belgesi: Union[int, str] = 0
    adli_sicil: Union[int, str] = 0
    ehliyet: Union[int, str] = 0
    kimlik_belgesi: Union[int, str] = 0
    fotograf: Union[int, str] = 0
    basvuru_formu: Union[int, str] = 0
    e_src_kaydi: Union[int, str] = 0
    alacagi_egitim: str = ""
    devam_egitimi: str = ""
    odeme_durumu: str = ""  
    tutar: Union[float, str] = 0
    tarih_1: str = ""
    odeme_1: Union[float, str] = 0
    tarih_2: str = ""
    odeme_2: Union[float, str] = 0
    tarih_3: str = ""
    odeme_3: Union[float, str] = 0
    tarih_4: str = ""
    odeme_4: Union[float, str] = 0
    tarih_5: str = ""
    odeme_5: Union[float, str] = 0
    tarih_6: str = ""
    odeme_6: Union[float, str] = 0
    kalan: Union[float, str] = 0
    aciklama: str = ""
    evrak_kayit_tarihi: str = ""
    inaktif: Union[int, str] = 0

class SinifEkleRequest(BaseModel):
    sinif_isim: str
    kursiyer_sayi: int = 0
    kursiyer_list: list = []

# Kullanıcı doğrulama
def authenticate_user(username: str, password: str):
    conn = get_db_connection()
    if not conn:
        return False
    
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kullanici WHERE kullanici_adi = ?", (username,))
    user = cursor.fetchone()
    conn.close()
    
    if not user:
        return False
    if not pwd_context.verify(password, user["sifre_hash"]):
        return False
    return user

# JWT token oluşturma
def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# Token doğrulama
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

async def get_current_user(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except jwt.PyJWTError:
        raise credentials_exception
    
    conn = get_db_connection()
    if not conn:
        raise credentials_exception
    
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kullanici WHERE kullanici_adi = ?", (username,))
    user = cursor.fetchone()
    conn.close()
    
    if user is None:
        raise credentials_exception
    return user

# Check bcrypt version compatibility for passlib
if not hasattr(bcrypt, "__about__"):
    import sys
    print(
        "WARNING: Incompatible bcrypt version detected. "
        "Please install 'bcrypt<4.0.0' for passlib compatibility.\n"
        "Run: pip install 'bcrypt<4.0.0'"
    )
    sys.exit(1)

# Endpoint'ler
@app.post("/token", response_model=Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends()):
    user = authenticate_user(form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["kullanici_adi"]}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/kullanici/", response_model=List[User])
async def list_kullanici(current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, kullanici_adi, rol FROM kullanici")
        rows = cursor.fetchall()
        users = [
            {"id": row["id"], "username": row["kullanici_adi"], "role": row["rol"]}
            for row in rows
        ]
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Kullanıcılar alınamadı: {str(e)}")
    conn.close()
    return users

@app.get("/kullanici/ayarlar")
async def kullanici_ayarlar(current_user: dict = Depends(get_current_user)):
    # Örnek veri, ihtiyaca göre düzenleyin
    return {
        "username": current_user["kullanici_adi"],
        "role": current_user["rol"],
        # ...diğer ayarlar...
    }

@app.post("/kullanici/sifre-degistir")
async def change_password(
    req: PasswordChangeRequest,
    current_user: dict = Depends(get_current_user)
):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    # Eski şifreyi doğrula
    if not pwd_context.verify(req.old_password, current_user["sifre_hash"]):
        conn.close()
        raise HTTPException(status_code=400, detail="Eski şifre yanlış")
    # Yeni şifreyi hashle ve güncelle
    new_hash = pwd_context.hash(req.new_password)
    cursor.execute(
        "UPDATE kullanici SET sifre_hash = ? WHERE kullanici_adi = ?",
        (new_hash, current_user["kullanici_adi"])
    )
    conn.commit()
    conn.close()
    return {"detail": "Şifre başarıyla değiştirildi."}

@app.post("/kullanici/ekle")
async def kullanici_ekle(
    req: UserCreateRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kullanici WHERE kullanici_adi = ?", (req.username,))
    if cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Kullanıcı zaten mevcut")
    sifre_hash = pwd_context.hash(req.password)
    cursor.execute(
        "INSERT INTO kullanici (kullanici_adi, sifre_hash, rol) VALUES (?, ?, ?)",
        (req.username, sifre_hash, req.role)
    )
    conn.commit()
    conn.close()
    return {"detail": "Kullanıcı başarıyla eklendi."}

@app.post("/kursiyer/ekle")
async def kursiyer_ekle(
    req: KursiyerEkleRequest,
    current_user: dict = Depends(get_current_user)
):
    # odeme_durumu alanına ödenecek tutarı yaz
    req_dict = req.dict()
    req_dict["odeme_durumu"] = str(req_dict.get("tutar", "")) if req_dict.get("tutar", "") != "" else ""

    # Convert numeric fields to correct types or default to 0
    def to_float(val):
        try:
            return float(val)
        except Exception:
            return 0.0

    def to_int(val):
        try:
            return int(val)
        except Exception:
            return 0

    for key in [
        "ogrenim_belgesi", "adres_belgesi", "adli_sicil", "ehliyet", "kimlik_belgesi",
        "fotograf", "basvuru_formu", "e_src_kaydi", "inaktif"
    ]:
        req_dict[key] = to_int(req_dict.get(key, 0))

    for key in [
        "tutar", "odeme_1", "odeme_2", "odeme_3", "odeme_4", "odeme_5", "odeme_6", "kalan"
    ]:
        req_dict[key] = to_float(req_dict.get(key, 0))

    conn = get_db_connection()
    if not conn:
        print("Kursiyer ekleme başarısız: Veritabanı bağlantı hatası")
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO kursiyer (
                kurs_durumu, aday_ismi, tel_no, tel_yakin, tc_kimlik,
                ogrenim_belgesi, adres_belgesi, adli_sicil, ehliyet, kimlik_belgesi,
                fotograf, basvuru_formu, e_src_kaydi, alacagi_egitim, devam_egitimi,
                odeme_durumu, tutar, tarih_1, odeme_1, tarih_2, odeme_2, tarih_3, odeme_3,
                tarih_4, odeme_4, tarih_5, odeme_5, tarih_6, odeme_6, kalan, aciklama,
                evrak_kayit_tarihi, inaktif
            ) VALUES (
                ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?
            )
        """, (
            req_dict.get("kurs_durumu", ""),
            req_dict.get("aday_ismi", ""),
            req_dict.get("tel_no", ""),
            req_dict.get("tel_yakin", ""),
            req_dict.get("tc_kimlik", ""),
            req_dict.get("ogrenim_belgesi", 0),
            req_dict.get("adres_belgesi", 0),
            req_dict.get("adli_sicil", 0),
            req_dict.get("ehliyet", 0),
            req_dict.get("kimlik_belgesi", 0),
            req_dict.get("fotograf", 0),
            req_dict.get("basvuru_formu", 0),
            req_dict.get("e_src_kaydi", 0),
            req_dict.get("alacagi_egitim", ""),
            req_dict.get("devam_egitimi", ""),
            req_dict.get("odeme_durumu", ""),
            req_dict.get("tutar", 0),
            req_dict.get("tarih_1", ""),
            req_dict.get("odeme_1", 0),
            req_dict.get("tarih_2", ""),
            req_dict.get("odeme_2", 0),
            req_dict.get("tarih_3", ""),
            req_dict.get("odeme_3", 0),
            req_dict.get("tarih_4", ""),
            req_dict.get("odeme_4", 0),
            req_dict.get("tarih_5", ""),
            req_dict.get("odeme_5", 0),
            req_dict.get("tarih_6", ""),
            req_dict.get("odeme_6", 0),
            req_dict.get("kalan", 0),
            req_dict.get("aciklama", ""),
            req_dict.get("evrak_kayit_tarihi", ""),
            req_dict.get("inaktif", 0)
        ))
        conn.commit()
        print(f"Kursiyer başarıyla eklendi: {req_dict.get('aday_ismi', '')} ({req_dict.get('tc_kimlik', '')})")
    except Exception as e:
        print(f"Kursiyer ekleme hatası: {str(e)} - Data: {req_dict}")
        conn.close()
        raise HTTPException(status_code=400, detail=f"Kursiyer eklenemedi: {str(e)}")
    conn.close()
    return {"detail": "Kursiyer başarıyla eklendi."}

@app.get("/kursiyer/liste")
async def kursiyer_liste(current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM kursiyer")
        rows = cursor.fetchall()
        kursiyerler = [dict(row) for row in rows]
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Kursiyerler alınamadı: {str(e)}")
    conn.close()
    return kursiyerler

@app.put("/kursiyer/guncelle/{kid}")
async def kursiyer_guncelle(kid: str, req: dict, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    field = "id"
    cursor.execute("PRAGMA table_info(kursiyer)")
    columns = [col[1] for col in cursor.fetchall()]
    if "id" not in columns:
        field = "tc_kimlik"
    allowed = set(columns) - {"id"}
    update_fields = []
    values = []

    int_fields = {"ogrenim_belgesi", "adres_belgesi", "adli_sicil", "ehliyet", "kimlik_belgesi",
                  "fotograf", "basvuru_formu", "e_src_kaydi", "inaktif"}
    float_fields = {"tutar", "odeme_1", "odeme_2", "odeme_3", "odeme_4", "odeme_5", "odeme_6", "kalan"}

    cursor.execute(f"SELECT * FROM kursiyer WHERE {field} = ?", (kid,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Kursiyer bulunamadı")
    # ...existing code...

    # Log için değişiklikleri topla
    log_degisiklikler = []
    log_odeme_tarih_alanlari = [
        "odeme_1", "odeme_2", "odeme_3", "odeme_4", "odeme_5", "odeme_6",
        "tarih_1", "tarih_2", "tarih_3", "tarih_4", "tarih_5", "tarih_6"
    ]
    is_user = current_user["rol"] == "user"

    for k in allowed:
        if k in req:
            v = req[k]
            if k in int_fields:
                try:
                    v = int(v)
                except Exception:
                    v = 0
            elif k in float_fields:
                try:
                    v = float(v)
                except Exception:
                    v = 0.0
            # Sadece user rolü ödeme/tarih alanlarını değiştirebilir, diğerlerini atla
            if is_user and k not in log_odeme_tarih_alanlari:
                continue
            # Değişiklik var mı kontrolü
            eski = row[k]
            if str(eski) != str(v):
                update_fields.append(f"{k} = ?")
                values.append(v)
                if is_user and k in log_odeme_tarih_alanlari:
                    log_degisiklikler.append(k)
            elif not is_user:
                update_fields.append(f"{k} = ?")
                values.append(v)
    if not update_fields:
        conn.close()
        raise HTTPException(
            status_code=400,
            detail=f"Güncellenecek veri yok veya alanlar hatalı. Gönderilen alanlar: {list(req.keys())}, izin verilenler: {list(allowed)}"
        )
    values.append(kid)
    try:
        cursor.execute(
            f"UPDATE kursiyer SET {', '.join(update_fields)} WHERE {field} = ?",
            values
        )
        # Log kaydı ekle (sadece user rolü ve ödeme/tarih alanı değiştiyse)
        if is_user and log_degisiklikler:
            aday_ismi = row["aday_ismi"]
            degisiklik = ", ".join(log_degisiklikler) + " değiştirildi"
            tarih = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cursor.execute(
                "INSERT INTO log (user, aday_ismi, degisiklik, tarih) VALUES (?, ?, ?, ?)",
                (current_user["kullanici_adi"], aday_ismi, degisiklik, tarih)
            )
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Güncelleme hatası: {str(e)}")
    conn.close()
    return {"detail": "Kursiyer başarıyla güncellendi."}

@app.delete("/kullanici/{user_id}/")
async def delete_kullanici(user_id: int, current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    if current_user["id"] == user_id:
        raise HTTPException(status_code=400, detail="Kendi hesabınızı silemezsiniz")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kullanici WHERE id = ?", (user_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    cursor.execute("DELETE FROM kullanici WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return {"detail": "Kullanıcı başarıyla silindi."}

@app.get("/kursiyer/ara")
async def kursiyer_ara(
    type: str = Query(..., regex="^(isim|tc)$"),
    value: str = Query(...),
    current_user: dict = Depends(get_current_user)
):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        if type == "isim":
            cursor.execute("SELECT * FROM kursiyer WHERE aday_ismi LIKE ?", (f"%{value}%",))
        elif type == "tc":
            cursor.execute("SELECT * FROM kursiyer WHERE tc_kimlik LIKE ?", (f"%{value}%",))
        else:
            conn.close()
            raise HTTPException(status_code=400, detail="Geçersiz arama türü")
        rows = cursor.fetchall()
        kursiyerler = [dict(row) for row in rows]
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Kursiyer arama hatası: {str(e)}")
    conn.close()
    return kursiyerler

@app.post("/api/sinif")
async def sinif_ekle(
    req: SinifEkleRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO sinif (sinif_isim, kursiyer_sayi, kursiyer_list) VALUES (?, ?, ?)",
            (
                req.sinif_isim,
                req.kursiyer_sayi,
                str(req.kursiyer_list),  
            )
        )
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Sınıf eklenemedi: {str(e)}")
    conn.close()
    return {"detail": "Sınıf başarıyla eklendi."}

@app.get("/api/siniflar")
async def sinif_listele(current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM sinif")
        rows = cursor.fetchall()
        siniflar = [dict(row) for row in rows]
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Sınıflar alınamadı: {str(e)}")
    conn.close()
    return siniflar

@app.delete("/api/sinif/{sinif_id}")
async def sinif_sil(sinif_id: int, current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM sinif WHERE id = ?", (sinif_id,))
    sinif = cursor.fetchone()
    if not sinif:
        conn.close()
        raise HTTPException(status_code=404, detail="Sınıf bulunamadı")
    cursor.execute("DELETE FROM sinif WHERE id = ?", (sinif_id,))
    conn.commit()
    conn.close()
    return {"detail": "Sınıf başarıyla silindi."}

@app.get("/api/belge/{kursiyer_id}")
async def belge_getir(kursiyer_id: str, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM belge WHERE kursiyer_id = ?", (kursiyer_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return {}
    return dict(row)

@app.post("/api/belge/yukle/{kursiyer_id}")
async def belge_yukle(
    kursiyer_id: str,
    belge_tipi: str = Form(...),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    allowed_keys = [
        "foto", "ogrenim_belgesi", "adres_belgesi", "adli_sicil",
        "ehliyet", "kimlik_belgesi", "basvuru_formu"
    ]
    # Dosya uzantısı veya tipi kontrolü yapılmaz, her şey kabul edilir
    # Kursiyer bilgisi çek
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kursiyer WHERE id = ? OR tc_kimlik = ?", (kursiyer_id, kursiyer_id))
    kursiyer = cursor.fetchone()
    if not kursiyer:
        conn.close()
        raise HTTPException(status_code=404, detail="Kursiyer bulunamadı")
    # Klasör adı: "aday_isim - alacagi_egitim" (boşluklar korunur)
    aday_ismi = (kursiyer["aday_ismi"] or "").strip()
    alacagi_egitim = (kursiyer["alacagi_egitim"] or "").strip().lower()
    uploads_dir = os.path.join(os.getcwd(), "uploads")
    klasor_adi = f"{aday_ismi} - {alacagi_egitim}".strip(" -")
    klasor_path = os.path.join(uploads_dir, klasor_adi)
    os.makedirs(klasor_path, exist_ok=True)
    ext = os.path.splitext(file.filename)[1]
    # Belge tipi 'foto' ise dosya adı 'FOTOGRAF' olarak büyük harflerle kaydedilsin
    if belge_tipi == "foto":
        dosya_adi = f"FOTOGRAF{ext}"
    else:
        dosya_adi = f"{belge_tipi}{ext}"
    dosya_path = os.path.join(klasor_path, dosya_adi)
    with open(dosya_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    rel_path = os.path.relpath(dosya_path, os.getcwd())
    kursiyer_real_id = kursiyer["id"]
    cursor.execute("SELECT * FROM belge WHERE kursiyer_id = ?", (kursiyer_real_id,))
    mevcut = cursor.fetchone()
    if mevcut:
        cursor.execute(
            f"UPDATE belge SET {belge_tipi} = ? WHERE kursiyer_id = ?",
            (rel_path, kursiyer_real_id)
        )
    else:
        fields = ["kursiyer_id"] + allowed_keys
        values = [kursiyer_real_id] + [rel_path if k == belge_tipi else None for k in allowed_keys]
        placeholders = ",".join(["?"] * len(fields))
        cursor.execute(
            f"INSERT INTO belge ({','.join(fields)}) VALUES ({placeholders})",
            values
        )
    conn.commit()
    cursor.execute("SELECT * FROM belge WHERE kursiyer_id = ?", (kursiyer_real_id,))
    belge_row = cursor.fetchone()
    conn.close()
    return dict(belge_row) if belge_row else {}

@app.post("/yedek/kursiyer-yukle")
async def yedek_kursiyer_yukle(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    # Save uploaded file to a temp file
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    errors = []
    imported = 0
    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active
        headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # Get kursiyer table columns
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(kursiyer)")
        kursiyer_columns = [col[1] for col in cursor.fetchall()]
        # Only use columns that exist in kursiyer table
        valid_headers = [h for h in headers if h in kursiyer_columns]
        if not valid_headers:
            raise Exception("Excel sütun başlıkları kursiyer tablosu ile eşleşmiyor.")
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_data = {h: (cell.value if cell.value is not None else "") for h, cell in zip(headers, row)}
            # Sadece aday_ismi doluysa ekle
            aday_ismi = row_data.get("aday_ismi", "")
            if not aday_ismi or str(aday_ismi).strip() == "":
                continue
            # Only keep valid columns except 'kalan'
            insert_data = {k: row_data[k] for k in valid_headers if k != "kalan"}
            # Telefon ve TC kimlik için değerleri daima string olarak al
            for key in ["tel_no", "tel_yakin", "tc_kimlik"]:
                if key in insert_data:
                    val = insert_data[key]
                    if val is None:
                        val = ""
                    # Excel'den gelen değer sayıysa stringe çevir
                    if isinstance(val, (int, float)):
                        # Ondalık kısmı yoksa tam sayı gibi göster
                        if isinstance(val, float) and val.is_integer():
                            val = str(int(val))
                        else:
                            val = str(val)
                    else:
                        val = str(val)
                    insert_data[key] = val
            # Kalanı hesapla
            try:
                tutar = float(row_data.get("tutar", 0) or 0)
            except Exception:
                tutar = 0.0
            toplam_odeme = 0.0
            for odeme_key in ["odeme_1", "odeme_2", "odeme_3", "odeme_4", "odeme_5", "odeme_6"]:
                try:
                    toplam_odeme += float(row_data.get(odeme_key, 0) or 0)
                except Exception:
                    pass
            kalan = round(tutar - toplam_odeme, 2)
            # Eğer kursiyer tablosunda 'kalan' sütunu varsa, insert_data'ya ekle
            if "kalan" in kursiyer_columns:
                insert_data["kalan"] = kalan
            # Try to insert
            try:
                insert_cols = list(insert_data.keys())
                placeholders = ",".join(["?"] * len(insert_cols))
                columns_str = ",".join(insert_cols)
                values = [insert_data[k] for k in insert_cols]
                cursor.execute(
                    f"INSERT INTO kursiyer ({columns_str}) VALUES ({placeholders})",
                    values
                )
                imported += 1
            except Exception as e:
                errors.append(f"{idx}. satır: {str(e)}")
        conn.commit()
        conn.close()
    except Exception as e:
        errors.append(f"Dosya okunamadı veya işlenemedi: {str(e)}")
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
    return {"imported": imported, "errors": errors}

@app.get("/yedek/db-indir")
async def yedek_db_indir(current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    db_path = os.path.join(os.getcwd(), "bandirma.db")
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail="Veritabanı dosyası bulunamadı")
    return FileResponse(
        db_path,
        media_type="application/octet-stream",
        filename="bandirma.db"
    )

@app.get("/api/sinav/{kursiyer_id}")
async def get_sinav(kursiyer_id: str, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM sinav WHERE kursiyer_id = ?", (kursiyer_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return {}
    return dict(row)

@app.post("/api/sinav/{kursiyer_id}")
async def upsert_sinav(
    kursiyer_id: str,
    req: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    # Sınav tablosu kolonları
    cursor.execute("PRAGMA table_info(sinav)")
    sinav_columns = [col[1] for col in cursor.fetchall()]
    # Sadece izin verilen alanlar
    allowed = set(sinav_columns) - {"id"}
    # Var mı kontrol et
    cursor.execute("SELECT * FROM sinav WHERE kursiyer_id = ?", (kursiyer_id,))
    mevcut = cursor.fetchone()
    # Sadece izin verilen alanları al
    data = {k: req[k] for k in allowed if k in req}
    data["kursiyer_id"] = int(kursiyer_id)

    yazili_keys = [f"sinav_{i+1}" for i in range(4)]
    uygulama_keys = [f"uygulama_{i+1}" for i in range(4)]

    # Yazılı sınav puanlarını 0-100 aralığında tut
    for k in yazili_keys:
        if k in data and data[k] is not None and data[k] != "":
            try:
                puan = float(data[k])
                if puan < 0:
                    puan = 0
                if puan > 100:
                    puan = 100
                data[k] = int(puan)
            except Exception:
                data[k] = 0

    # Eğer yazılı sınavlardan herhangi biri >= 70 değilse uygulama sınavlarını "" yap
    yazili_gecti = any(
        (data.get(k) is not None and data.get(k) != "" and float(data.get(k)) >= 70)
        for k in yazili_keys
    )
    if not yazili_gecti:
        for k in uygulama_keys:
            data[k] = ""
    else:
        # Uygulama sınavı değerlerini None ise "" yap
        for k in uygulama_keys:
            if data.get(k) is None:
                data[k] = ""

    try:
        if mevcut:
            update_fields = [f"{k} = ?" for k in data if k != "kursiyer_id"]
            values = [data[k] for k in data if k != "kursiyer_id"]
            values.append(kursiyer_id)
            cursor.execute(
                f"UPDATE sinav SET {', '.join(update_fields)} WHERE kursiyer_id = ?",
                values
            )
        else:
            insert_cols = list(data.keys())
            placeholders = ",".join(["?"] * len(insert_cols))
            columns_str = ",".join(insert_cols)
            values = [data[k] for k in insert_cols]
            cursor.execute(
                f"INSERT INTO sinav ({columns_str}) VALUES ({placeholders})",
                values
            )
        # Uygulama sınavlarından biri "Geçti" ise kursiyer.inaktif = 1 yap
        uygulama_gecti = any(
            (data.get(f"uygulama_{i+1}", "") == "Geçti") for i in range(4)
        )
        if uygulama_gecti:
            cursor.execute("UPDATE kursiyer SET inaktif = 1 WHERE id = ?", (kursiyer_id,))
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Sınav kaydı hatası: {str(e)}")
    conn.close()
    return {"detail": "Sınav bilgisi kaydedildi."}

@app.get("/api/sinavlar/filtreli")
async def sinavlar_filtreli(
    filtre: str = Query(..., regex="^(yazilidan_kalanlar|sadece_yazilidan_gecenler|her_ikisinden_kalanlar|kursu_tamamlayanlar)$"),
    current_user: dict = Depends(get_current_user)
):
    """
    Filtreye göre kursiyer ve sınav bilgilerini döndürür.
    Dönüş: [{aday_ismi, tel_no, tc_kimlik, alacagi_egitim, sinif, ...sinav alanları}]
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            k.id as kursiyer_id, k.aday_ismi, k.tel_no, k.tc_kimlik, k.alacagi_egitim, k.sinif,
            s.*
        FROM kursiyer k
        LEFT JOIN sinav s ON k.id = s.kursiyer_id
    """)
    rows = cursor.fetchall()
    result = []
    for row in rows:
        row = dict(row)
        # SINIFI OLMAYANLARI ATLAMAK İÇİN:
        if not row.get("sinif") or str(row.get("sinif")).strip() == "":
            continue
        yazili = [row.get(f"sinav_{i+1}") for i in range(4)]
        uygulama = [row.get(f"uygulama_{i+1}") for i in range(4)]
        yazili_gecen = any(
            (puan is not None and str(puan) != "" and float(puan) >= 70)
            for puan in yazili
        )
        yazili_kalan = all(
            (puan is None or str(puan) == "" or float(puan) < 70)
            for puan in yazili
        )
        uygulama_gecen = any(
            (puan == "Geçti")
            for puan in uygulama
        )
        uygulama_kalan = all(
            (puan != "Geçti")
            for puan in uygulama
        )
        kursu_tamamlayan = uygulama_gecen

        # Yeni filtre mantığı:
        if filtre == "yazilidan_kalanlar" and yazili_kalan:
            result.append(row)
        elif filtre == "sadece_yazilidan_gecenler" and yazili_gecen and uygulama_kalan:
            result.append(row)
        elif filtre == "her_ikisinden_kalanlar" and yazili_kalan and uygulama_kalan:
            result.append(row)
        elif filtre == "kursu_tamamlayanlar" and kursu_tamamlayan:
            result.append(row)
    conn.close()
    filtered = []
    for row in result:
        filtered.append({
            "aday_ismi": row.get("aday_ismi"),
            "tel_no": row.get("tel_no"),
            "tc_kimlik": row.get("tc_kimlik"),
            "alacagi_egitim": row.get("alacagi_egitim"),
            "sinif": row.get("sinif"),
            "sinav_1": row.get("sinav_1"),
            "sinav_2": row.get("sinav_2"),
            "sinav_3": row.get("sinav_3"),
            "sinav_4": row.get("sinav_4"),
            "uygulama_1": row.get("uygulama_1"),
            "uygulama_2": row.get("uygulama_2"),
            "uygulama_3": row.get("uygulama_3"),
            "uygulama_4": row.get("uygulama_4"),
        })
    return filtered

@app.get("/api/loglar")
async def loglari_getir(current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT l.user, l.aday_ismi, l.degisiklik, l.tarih
            FROM log l
            ORDER BY l.tarih DESC
        """)
        rows = cursor.fetchall()
        loglar = [dict(row) for row in rows]
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Loglar alınamadı: {str(e)}")
    conn.close()
    return loglar

@app.delete("/api/loglar/temizle")
async def loglari_temizle(current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM log")
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Loglar silinemedi: {str(e)}")
    conn.close()
    return {"detail": "Tüm loglar silindi."}

# Test endpoint
@app.get("/")
async def root():
    return {"message": "Bandırma CRM API çalışıyor"}

@app.exception_handler(404)
async def not_found_handler(request: Request, exc):
    return JSONResponse(
        status_code=404,
        content={"detail": f"Endpoint {request.url.path} not found."}
    )

@app.get("/users/me/", response_model=User)
async def read_users_me(current_user: dict = Depends(get_current_user)):
    return {
        "id": current_user["id"],
        "username": current_user["kullanici_adi"],
        "role": current_user["rol"]
    }